"""
excel_read — read a filled review .xlsx (from excel_template.py) back into an approved plan.

The operator's edits in Excel are mapped back onto the Line contract:
  * Claim = No        -> claimGuess "personal" (excluded from the claim)
  * ExpenseType       -> proposed.typeCode (display mapped back to its code) + typeDisplay
  * Attendees text    -> proposed.attendees [{name, company:"", title:""}]
  * Split5050 = Yes   -> proposed.split = True

The display -> code map depends on which entity you are filing under: the AU chart
(default) or the HK one (schema/hk_expense_types.json). The chart comes from, in order:

  1. the workbook's own `Header` sheet, if it has a label/value row `Chart | hk`
     (this is what Waypoint exports);
  2. an explicit `--chart`;
  3. `au`.

If (1) and (2) are both present and DISAGREE the read fails naming both values — one of
them is wrong and guessing is how a wrong claim gets filed. Two display strings exist in
BOTH charts with different codes ("Meals & Ent'mnt - Client" = AU MEALCLI / HK MEAL50,
"Employee Relocation" = AU EMPRELO / HK EMPRELC), so the wrong chart mis-files the most
common line in a pack without tripping any unknown-type check.

An ExpenseType the chosen chart doesn't know is an ERROR, not a blank: dropping the type
off a real claim line is an audit problem. Pass --allow-unknown-types to fall back to
the old lenient behaviour (typeCode left null).

Usage:
    python tools/excel_read.py review.xlsx --out approved.json
    python tools/excel_read.py review.xlsx --chart hk --out approved.json
"""
from __future__ import annotations

import argparse
import os
import sys
from typing import Optional

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cbre_lib import (CHART_NAMES, ExpenseChart, add_chart_argument,  # noqa: E402
                      dump_json, load_chart)

import openpyxl  # noqa: E402

HEADER_SHEET = "header"     # matched case-insensitively
CHART_LABEL = "chart"       # the label of the Header row that names the chart


class UnknownExpenseTypeError(ValueError):
    """A sheet holds ExpenseType text that the selected chart has no code for.

    `unknown` is [{"row": <sheet row>, "id": <line id>, "value": <display text>}, ...].
    """

    def __init__(self, unknown: list[dict], chart_name: str):
        self.unknown = unknown
        self.chart_name = chart_name
        detail = "; ".join(f"row {u['row']} ({u['id']}): {u['value']!r}" for u in unknown)
        super().__init__(
            f"{len(unknown)} ExpenseType value(s) are not in the '{chart_name}' chart: {detail}")


class ChartConflictError(ValueError):
    """--chart says one entity, the workbook's Header sheet says another."""

    def __init__(self, requested: str, declared: str):
        self.requested = requested
        self.declared = declared
        super().__init__(
            f"chart conflict: --chart {requested!r} was given, but the workbook's Header "
            f"sheet declares Chart {declared!r}. One of them is wrong - re-run with "
            f"--chart {declared} to accept the sheet, or fix the sheet's Chart row.")


class SheetChartError(ValueError):
    """The Header sheet names a chart that doesn't exist."""

    def __init__(self, declared: str):
        self.declared = declared
        super().__init__(
            f"the workbook's Header sheet declares Chart {declared!r}, which is not a known "
            f"chart ({', '.join(CHART_NAMES)}) - fix the sheet rather than guessing.")


def _find_sheet(wb, name: str):
    for sheet_name in wb.sheetnames:
        if str(sheet_name).strip().lower() == name:
            return wb[sheet_name]
    return None


def declared_chart(wb) -> Optional[str]:
    """The chart named by the workbook's `Header` sheet, or None if it doesn't say."""
    ws = _find_sheet(wb, HEADER_SHEET)
    if ws is None:
        return None
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        label = str(row[0] or "").strip().rstrip(":").lower()
        if label == CHART_LABEL:
            value = str(row[1] or "").strip() if len(row) > 1 else ""
            return value or None
    return None


def resolve_chart(wb, requested: Optional[str] = None) -> ExpenseChart:
    """Pick the chart for a workbook: its Header sheet wins, `requested` must not clash.

    `requested` is what the operator explicitly asked for (or None if they didn't).
    """
    declared = declared_chart(wb)
    if declared is None:
        return load_chart(requested or "au")
    key = declared.strip().lower()
    if key not in CHART_NAMES:
        raise SheetChartError(declared)
    if requested is not None and str(requested).strip().lower() != key:
        raise ChartConflictError(str(requested).strip().lower(), key)
    return load_chart(key)


def _yes(v) -> bool:
    return str(v).strip().lower() in ("yes", "y", "true", "1")


def _parse_attendees(text):
    """'Surname,First; Surname,First' -> [{name, company:'', title:''}]."""
    out = []
    for chunk in str(text or "").split(";"):
        name = chunk.strip()
        if name:
            out.append({"name": name, "company": "", "title": ""})
    return out


def _parse_flags(text):
    return [f.strip() for f in str(text or "").split(";") if f.strip()]


def _num(v):
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return v


def read_plan(path: str, chart: str | ExpenseChart | None = None,
              allow_unknown: bool = False) -> dict:
    """Read a filled review sheet into {"chart": <name>, "chartSource": …, "lines": [...]}.

    `chart` is the explicitly-requested chart, or None to take the workbook's own
    (Header sheet) chart and fall back to AU. `chartSource` records which of those it
    was ("header" | "flag" | "default") so a run is auditable after the fact. Raises
    ChartConflictError if both are given and disagree, SheetChartError if the sheet names
    an unknown chart, and UnknownExpenseTypeError if any non-blank ExpenseType is absent
    from the resolved chart (unless allow_unknown). A blank ExpenseType means "not set".
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    if isinstance(chart, ExpenseChart):
        source = "caller"
    elif declared_chart(wb) is not None:
        source = "header"
        chart = resolve_chart(wb, chart)
    else:
        source = "flag" if chart is not None else "default"
        chart = resolve_chart(wb, chart)
    display_to_code = chart.display_to_code

    ws = _find_sheet(wb, "review")
    if ws is None:
        # No Review sheet: fall back to the first sheet that isn't the Header block.
        ws = next((wb[n] for n in wb.sheetnames
                   if str(n).strip().lower() != HEADER_SHEET), wb.active)

    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def get(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    lines = []
    unknown: list[dict] = []
    for rownum, row in enumerate(rows, start=2):   # header was row 1
        if get(row, "ID") in (None, ""):
            continue  # skip blank trailing rows

        line_id = str(get(row, "ID"))
        claim_yes = _yes(get(row, "Claim"))
        display = (get(row, "ExpenseType") or "")
        display = str(display).strip()
        code = display_to_code.get(display)
        if display and code is None:
            unknown.append({"row": rownum, "id": line_id, "value": display})
        split = _yes(get(row, "Split5050"))
        attendees = _parse_attendees(get(row, "Attendees"))

        proposed = {
            "typeCode": code,
            "typeDisplay": display or None,
            "needsAttendees": code in chart.meal_types_need_attendees if code else False,
            "attendees": attendees,
            "split": split,
        }
        line = {
            "id": line_id,
            "date": get(row, "Date"),
            "merchant": get(row, "Merchant"),
            "amount": _num(get(row, "Amount")),
            "currency": get(row, "Currency"),
            "claimGuess": "business" if claim_yes else "personal",
            "proposed": proposed,
            "flags": _parse_flags(get(row, "Notes")),
        }
        lines.append(line)

    if unknown and not allow_unknown:
        raise UnknownExpenseTypeError(unknown, chart.name)
    return {"chart": chart.name, "chartSource": source, "lines": lines}


def read_workbook(path: str, chart: str | ExpenseChart | None = None,
                  allow_unknown: bool = False) -> list[dict]:
    """Just the plan lines — see read_plan() for the chart-resolution rules."""
    return read_plan(path, chart=chart, allow_unknown=allow_unknown)["lines"]


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Read a filled review .xlsx back into an approved plan JSON.")
    ap.add_argument("xlsx", help="filled review spreadsheet (from excel_template.py)")
    ap.add_argument("--out", required=True, help="output approved plan JSON path")
    add_chart_argument(ap, default=None)   # None = "not explicitly asked for"
    ap.add_argument("--allow-unknown-types", action="store_true",
                    help="don't fail on ExpenseType values missing from the chart "
                         "(leaves typeCode null — the old, silent behaviour)")
    args = ap.parse_args()

    try:
        plan = read_plan(args.xlsx, chart=args.chart,
                         allow_unknown=args.allow_unknown_types)
    except (ChartConflictError, SheetChartError) as exc:
        print(f"ERROR: {exc} Nothing was written.", file=sys.stderr)
        sys.exit(2)
    except UnknownExpenseTypeError as exc:
        print(f"ERROR: {len(exc.unknown)} row(s) have an ExpenseType that is not in the "
              f"'{exc.chart_name}' chart. Nothing was written.", file=sys.stderr)
        for u in exc.unknown:
            print(f"  row {u['row']}  {u['id']}  {u['value']!r}", file=sys.stderr)
        other = "hk" if exc.chart_name == "au" else "au"
        print(f"Fix the spelling in the sheet, or re-run with --chart {other} if the sheet "
              f"was written against the other entity's chart.", file=sys.stderr)
        sys.exit(2)

    dump_json(plan, args.out)
    lines = plan["lines"]
    claimed = sum(1 for ln in lines if ln["claimGuess"] != "personal")
    source = {"header": "declared by the sheet's Header",
              "flag": "from --chart",
              "default": "default"}.get(plan["chartSource"], plan["chartSource"])
    print(f"Read {len(lines)} lines ({claimed} claimed) using the "
          f"{plan['chart']} expense chart ({source}) -> {args.out}")


if __name__ == "__main__":
    main()
