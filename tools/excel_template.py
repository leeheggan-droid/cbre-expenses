"""
excel_template — turn a classified.json into a review/edit spreadsheet ("easy mode").

An operator opens the .xlsx, fixes the Claim / ExpenseType / Attendees / Split5050
columns with dropdowns instead of hand-editing JSON, then excel_read.py turns the
filled sheet back into an approved plan.

Usage:
    python tools/excel_template.py classified.json --out review.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from cbre_lib import EXPENSE_TYPES, load_json  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

# Column order on the review sheet. (header, width)
COLUMNS = [
    ("ID", 10),
    ("Date", 12),
    ("Merchant", 34),
    ("Amount", 14),
    ("Currency", 10),
    ("Claim", 9),
    ("ExpenseType", 30),
    ("Attendees", 40),
    ("Split5050", 11),
    ("Notes", 60),
]


def _attendees_text(attendees) -> str:
    """Render proposed.attendees as 'Surname,First; Surname,First' (names only)."""
    names = [str(a.get("name", "")).strip() for a in (attendees or []) if a.get("name")]
    return "; ".join(names)


def build_workbook(classified: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    # Hidden helper sheet holds the dropdown source lists (the ExpenseType list is
    # too long for an inline DataValidation formula, which caps at 255 chars).
    lists = wb.create_sheet("Lists")
    type_displays = list(EXPENSE_TYPES.values())
    for i, disp in enumerate(type_displays, start=1):
        lists.cell(row=i, column=1, value=disp)
    lists.cell(row=1, column=2, value="Yes")
    lists.cell(row=2, column=2, value="No")
    lists.sheet_state = "hidden"

    # Header row.
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for c, (name, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = width

    col = {name: i for i, (name, _) in enumerate(COLUMNS, start=1)}

    lines = classified.get("lines", [])
    for r, ln in enumerate(lines, start=2):
        proposed = ln.get("proposed") or {}
        ws.cell(row=r, column=col["ID"], value=ln.get("id"))
        ws.cell(row=r, column=col["Date"], value=ln.get("date"))
        ws.cell(row=r, column=col["Merchant"], value=ln.get("merchant"))
        amt = ws.cell(row=r, column=col["Amount"], value=ln.get("amount"))
        amt.number_format = "#,##0.00"
        ws.cell(row=r, column=col["Currency"], value=ln.get("currency"))
        ws.cell(row=r, column=col["Claim"],
                value="Yes" if ln.get("claimGuess") != "personal" else "No")
        ws.cell(row=r, column=col["ExpenseType"], value=proposed.get("typeDisplay"))
        ws.cell(row=r, column=col["Attendees"],
                value=_attendees_text(proposed.get("attendees")))
        ws.cell(row=r, column=col["Split5050"],
                value="Yes" if proposed.get("split") else "No")
        notes = ws.cell(row=r, column=col["Notes"], value="; ".join(ln.get("flags") or []))
        notes.alignment = Alignment(wrap_text=False, vertical="center")

    last_row = max(len(lines) + 1, 2)

    # Dropdowns (skip the header row).
    yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    yesno.error = "Choose Yes or No."
    yesno.prompt = "Choose Yes or No."
    ws.add_data_validation(yesno)
    yesno.add(f"{get_column_letter(col['Claim'])}2:{get_column_letter(col['Claim'])}{last_row}")
    yesno.add(f"{get_column_letter(col['Split5050'])}2:{get_column_letter(col['Split5050'])}{last_row}")

    type_ref = f"Lists!$A$1:$A${len(type_displays)}"
    types_dv = DataValidation(type="list", formula1=type_ref, allow_blank=True)
    types_dv.error = "Pick an expense type from the list."
    types_dv.prompt = "Pick the CBRE expense type."
    ws.add_data_validation(types_dv)
    types_dv.add(f"{get_column_letter(col['ExpenseType'])}2:{get_column_letter(col['ExpenseType'])}{last_row}")

    # Freeze the header row so it stays visible while scrolling.
    ws.freeze_panes = "A2"
    return wb


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Write a review/edit .xlsx from a classified.json (Excel easy mode).")
    ap.add_argument("classified", help="classified JSON (from classify.py)")
    ap.add_argument("--out", required=True, help="output .xlsx path")
    args = ap.parse_args()

    classified = load_json(args.classified)
    wb = build_workbook(classified)
    wb.save(args.out)
    print(f"Wrote {len(classified.get('lines', []))} lines -> {args.out}")


if __name__ == "__main__":
    main()
