"""Round-trip tests for the Excel "easy mode": classified.json -> .xlsx -> approved.json.

Run with pytest, or standalone:  python tests/test_excel.py
"""
import os
import subprocess
import sys
import tempfile

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TOOLS = os.path.join(ROOT, "tools")
SAMPLES = os.path.join(ROOT, "samples")
sys.path.insert(0, TOOLS)

import openpyxl  # noqa: E402

import cbre_lib as L  # noqa: E402
import excel_template as ET  # noqa: E402
import excel_read as ER  # noqa: E402

_TMP = tempfile.mkdtemp(prefix="cbre_excel_test_")


def _make_classified():
    """Generate a classified.json from the sample data via the real pipeline."""
    lines = os.path.join(_TMP, "lines.json")
    classified = os.path.join(_TMP, "classified.json")
    subprocess.run(
        [sys.executable, os.path.join(TOOLS, "parse_statement.py"),
         os.path.join(SAMPLES, "bank-sample.csv"), "--out", lines],
        check=True, capture_output=True)
    subprocess.run(
        [sys.executable, os.path.join(TOOLS, "classify.py"), lines,
         "--run-config", os.path.join(SAMPLES, "run-config.example.json"),
         "--roster", os.path.join(SAMPLES, "roster.example.json"),
         "--out", classified],
        check=True, capture_output=True)
    return L.load_json(classified)


def test_template_builds_with_dropdowns_and_prefill():
    classified = _make_classified()
    wb = ET.build_workbook(classified)
    ws = wb["Review"]

    # Header row present + frozen.
    headers = [c.value for c in ws[1]]
    for col in ("ID", "Date", "Merchant", "Amount", "Currency", "Claim",
                "ExpenseType", "Attendees", "Split5050", "Notes"):
        assert col in headers, f"missing column {col}"
    assert ws.freeze_panes == "A2"

    # Lists helper sheet is hidden and present.
    assert "Lists" in wb.sheetnames
    assert wb["Lists"].sheet_state == "hidden"
    # Two data validations: Yes/No and the expense-type list.
    assert len(ws.data_validations.dataValidation) == 2

    # Prefill: first data line is the Uber/taxi, claimGuess business -> Claim "Yes".
    idx = {h: i for i, h in enumerate(headers)}
    row2 = [c.value for c in ws[2]]
    assert row2[idx["ID"]] == "L001"
    assert row2[idx["Claim"]] == "Yes"
    assert row2[idx["ExpenseType"]] == "Taxis - Business Use"


def test_attendees_prefill_format():
    classified = _make_classified()
    wb = ET.build_workbook(classified)
    ws = wb["Review"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    # L002 is the client meal with two roster attendees.
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["ID"]] == "L002":
            assert row[idx["Attendees"]] == "Smith,John; Doe,Jane"
            assert row[idx["Split5050"]] == "Yes"
            break
    else:
        raise AssertionError("L002 row not found")


def test_round_trip_edits_survive():
    classified = _make_classified()
    xlsx = os.path.join(_TMP, "review.xlsx")
    approved = os.path.join(_TMP, "approved.json")

    wb = ET.build_workbook(classified)
    wb.save(xlsx)

    # Programmatically fill/edit a couple of cells, like an operator would.
    wb2 = openpyxl.load_workbook(xlsx)
    ws = wb2["Review"]
    headers = [c.value for c in ws[1]]
    col = {h: i + 1 for i, h in enumerate(headers)}

    rows_by_id = {}
    for r in range(2, ws.max_row + 1):
        rows_by_id[ws.cell(row=r, column=col["ID"]).value] = r

    # Edit 1: mark L007 (unknown TRAVOTH) as NOT claimed.
    r7 = rows_by_id["L007"]
    ws.cell(row=r7, column=col["Claim"], value="No")

    # Edit 2: change L001's expense type display + set its attendees + split.
    r1 = rows_by_id["L001"]
    ws.cell(row=r1, column=col["ExpenseType"], value="Meals & Ent'mnt - Client")
    ws.cell(row=r1, column=col["Attendees"], value="Brown,Alice; Green,Bob")
    ws.cell(row=r1, column=col["Split5050"], value="Yes")
    wb2.save(xlsx)

    # Read back.
    ER_out = ER.read_workbook(xlsx)
    by_id = {ln["id"]: ln for ln in ER_out}

    # Edit 1 survived: excluded as personal.
    assert by_id["L007"]["claimGuess"] == "personal"

    # Edit 2 survived: display mapped back to its code, attendees parsed, split set.
    l1 = by_id["L001"]
    assert l1["claimGuess"] == "business"
    assert l1["proposed"]["typeDisplay"] == "Meals & Ent'mnt - Client"
    assert l1["proposed"]["typeCode"] == "MEALCLI"          # display -> code mapping
    assert l1["proposed"]["needsAttendees"] is True         # MEALCLI needs attendees
    assert l1["proposed"]["split"] is True
    names = [a["name"] for a in l1["proposed"]["attendees"]]
    assert names == ["Brown,Alice", "Green,Bob"]
    assert all(a["company"] == "" and a["title"] == "" for a in l1["proposed"]["attendees"])


def test_unedited_round_trip_preserves_type_codes():
    classified = _make_classified()
    xlsx = os.path.join(_TMP, "review2.xlsx")
    ET.build_workbook(classified).save(xlsx)
    out = ER.read_workbook(xlsx)
    by_id = {ln["id"]: ln for ln in out}
    # Every prefilled display must map back to the original proposed code.
    for ln in classified["lines"]:
        if ln["claimGuess"] == "personal":
            continue
        got = by_id[ln["id"]]["proposed"]["typeCode"]
        assert got == ln["proposed"]["typeCode"], f"{ln['id']}: {got} != {ln['proposed']['typeCode']}"


def _run_standalone():
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    for fn in fns:
        fn()
        print(f"PASS {fn.__name__}")
    print(f"\nAll {len(fns)} tests passed.")


if __name__ == "__main__":
    _run_standalone()
